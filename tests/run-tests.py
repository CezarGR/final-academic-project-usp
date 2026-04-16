#!/usr/bin/env python3
"""
Orquestra cenários de teste de carga: sobe stack (docker compose), prepara MongoDB,
valida portas, executa k6 e consolida métricas em result.xlsx.

Cenários: pastas ``scenario-*`` com ``config.json`` (``applicationUri``, ``runDockerCompose``,
``externalApplication``, ``prometheusUri`` quando app/Prometheus estão fora mas Grafana sobe no compose).
Script k6 único: ``tests/k6-load-test.js``.
CLI: ``--scenario PASTA``; ``--skip-scenario PASTA`` (repetir para vários) exclui da fila;
``--repeat N`` repete o fluxo completo.

Dados TSDB do Prometheus (cenários com serviço ``prometheus`` no compose) vão para um volume Docker
externo nomeado ``scenario_<n>_<rodada>_execution`` (ex.: ``scenario_2_2_execution`` na 2ª rodada
do ``scenario-2``). Para inspecionar depois: ``tests/docker-compose.metrics-view.yml`` com
``PROMETHEUS_DATA_VOLUME_NAME`` apontando para esse volume.

Dependências: pandas, openpyxl, requests, pymongo, playwright (+ ``playwright install chromium``);
Docker (compose v2 ou docker-compose). Capturas Grafana usam o JSON do dashboard para listar painéis.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
import pandas as pd
import requests
from bson import json_util
from openpyxl import load_workbook
from openpyxl.styles import Font
from pymongo import MongoClient

# Caminho padrão: diretório onde está este script
TESTS_ROOT = Path(__file__).resolve().parent
REPO_ROOT = TESTS_ROOT.parent
USERS_SEED_FILE = TESTS_ROOT / "users-collection.json"
OUTPUT_FILE = TESTS_ROOT / "result.xlsx"
DEFAULT_GRAFANA_DASHBOARD_JSON = (
    REPO_ROOT / "monitoring/grafana/dashboards/provisionamento-infra.json"
)


def grafana_dashboard_json_for_scenario(scenario_dir: Path) -> Path:
    """JSON usado pelo Grafana neste cenário (volume ./monitoring/...), senão o padrão do repositório."""
    local = scenario_dir / "monitoring/grafana/dashboards/provisionamento-infra.json"
    if local.is_file():
        return local
    return DEFAULT_GRAFANA_DASHBOARD_JSON

PROMETHEUS_URL = "http://localhost:9090"
GRAFANA_URL = os.environ.get("GRAFANA_URL", "http://localhost:3000").rstrip("/")

DEFAULT_PORT_CHECKS: tuple[tuple[str, int], ...] = (
    ("aplicação", 8080),
    ("prometheus", 9090),
    ("grafana", 3000),
)

GRAFANA_ONLY_PORT_CHECKS: tuple[tuple[str, int], ...] = (("grafana", 3000),)

GENERATED_PROMETHEUS_DS = "prometheus.generated.yml"

COMPOSE_SERVICE_WAIT_S = 120
PORT_RETRY_INTERVAL_S = 2

SKIP_SCENARIO_DIR_NAMES = frozenset({"scenario-pattern"})


@dataclass
class ScenarioConfig:
    databaseLocation: str
    scenario: str
    description: str
    cpus: float
    memory: int
    testFile: str
    applicationUri: str
    healthCheckUri: str | None
    runDockerCompose: bool
    externalApplication: bool
    prometheusUri: str | None


def discover_scenarios(tests_root: Path) -> list[Path]:
    """Pastas scenario-* com config.json; docker-compose obrigatório se runDockerCompose."""
    found: list[Path] = []
    for p in sorted(tests_root.glob("scenario-*")):
        if not p.is_dir():
            continue
        if p.name in SKIP_SCENARIO_DIR_NAMES:
            continue
        cfg_path = p / "config.json"
        if not cfg_path.is_file():
            continue
        try:
            with open(cfg_path, encoding="utf-8") as f:
                raw = json.load(f)
        except json.JSONDecodeError:
            continue
        run_dc = raw.get("runDockerCompose", True)
        if isinstance(run_dc, str):
            run_dc = run_dc.lower() in ("true", "1", "yes")
        if run_dc and not (p / "docker-compose.yml").is_file():
            continue
        found.append(p)
    return found


def format_datetime_local_from_epoch(epoch_s: int) -> str:
    """Converte epoch Unix (segundos) para string local ``YYYY-MM-DD HH:MM:SS``."""
    return datetime.fromtimestamp(int(epoch_s)).strftime("%Y-%m-%d %H:%M:%S")


def format_duration_hms(seconds: float) -> str:
    """Formata duração em horas, minutos e segundos (ex.: 1h 5m 3s)."""
    secs = max(0, int(round(seconds)))
    h, r = divmod(secs, 3600)
    m, s = divmod(r, 60)
    return f"{h}h {m}m {s}s"


def load_scenario_config(scenario_dir: Path) -> ScenarioConfig:
    with open(scenario_dir / "config.json", encoding="utf-8") as f:
        raw = json.load(f)
    loc = str(raw.get("databaseLocation", "internal")).lower()
    if loc not in ("internal", "external"):
        raise ValueError(f"databaseLocation inválido em {scenario_dir}: {loc}")
    app_uri = (raw.get("applicationUri") or raw.get("application_uri") or "").strip()
    if not app_uri:
        raise ValueError(f"applicationUri obrigatório em {scenario_dir / 'config.json'}")
    run_dc = raw.get("runDockerCompose", True)
    if isinstance(run_dc, str):
        run_dc = run_dc.lower() in ("true", "1", "yes")
    hc = raw.get("healthCheckUri") or raw.get("health_check_uri")
    hc = hc.strip() if isinstance(hc, str) and hc.strip() else None
    mem = raw.get("memory", 0)
    ext = raw.get("externalApplication", False)
    if isinstance(ext, str):
        ext = ext.lower() in ("true", "1", "yes")
    prom = raw.get("prometheusUri") or raw.get("prometheus_uri")
    prom = prom.strip() if isinstance(prom, str) and prom.strip() else None
    return ScenarioConfig(
        databaseLocation=loc,
        scenario=str(raw.get("scenario", "")),
        description=str(raw.get("description", "")),
        cpus=float(raw.get("cpus", 0)),
        memory=int(mem) if mem is not None else 0,
        testFile=str(raw.get("testFile", "k6-load-test.js")),
        applicationUri=app_uri,
        healthCheckUri=hc,
        runDockerCompose=bool(run_dc),
        externalApplication=bool(ext),
        prometheusUri=prom,
    )


def resolve_health_check_url(cfg: ScenarioConfig) -> str:
    """URL do /actuator/health acessível a partir do host que roda o Python."""
    if cfg.healthCheckUri:
        return cfg.healthCheckUri.strip()
    if cfg.runDockerCompose and not cfg.externalApplication:
        return "http://127.0.0.1:8080/actuator/health"
    base = cfg.applicationUri.rstrip("/")
    return f"{base}/actuator/health"


def port_checks_for_config(cfg: ScenarioConfig) -> tuple[tuple[str, int], ...]:
    """Com app externa e compose só Grafana, só a porta 3000 precisa estar aberta no host."""
    if cfg.externalApplication and cfg.runDockerCompose:
        return GRAFANA_ONLY_PORT_CHECKS
    return DEFAULT_PORT_CHECKS


def write_grafana_prometheus_datasource(scenario_dir: Path, prometheus_uri: str) -> Path:
    """Gera YAML de datasource para o Grafana apontar para Prometheus externo."""
    base = prometheus_uri.rstrip("/")
    out_dir = scenario_dir / "monitoring/grafana/provisioning/datasources"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / GENERATED_PROMETHEUS_DS
    path.write_text(
        "apiVersion: 1\n\ndatasources:\n"
        "  - name: Prometheus\n"
        "    type: prometheus\n"
        "    access: proxy\n"
        f"    url: {base}\n"
        "    isDefault: true\n",
        encoding="utf-8",
    )
    print(f"Datasource Grafana → Prometheus gravado em {path}")
    return path


def parse_dotenv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path.is_file():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" in line:
            k, _, v = line.partition("=")
            out[k.strip()] = v.strip().strip('"').strip("'")
    return out


def database_name_from_uri(uri: str) -> str:
    """Extrai o database do path da URI sem parse_uri (evita DNS em mongodb+srv)."""
    uri = uri.split("?", 1)[0]
    for prefix in ("mongodb+srv://", "mongodb://"):
        if uri.startswith(prefix):
            rest = uri[len(prefix) :]
            break
    else:
        rest = uri
    if "/" not in rest:
        return "transactiondb"
    path = rest.split("/", 1)[1]
    db = path.split("/")[0].strip()
    if not db:
        return "transactiondb"
    return db


def mongo_uri_for_scenario(scenario_dir: Path, cfg: ScenarioConfig) -> str:
    if cfg.databaseLocation == "internal":
        return "mongodb://127.0.0.1:27017/transactiondb"
    env_map = parse_dotenv(scenario_dir / ".env")
    uri = env_map.get("SPRING_DATA_MONGO_DB_URI")
    if not uri:
        raise RuntimeError(
            f"databaseLocation=external exige SPRING_DATA_MONGO_DB_URI em {scenario_dir / '.env'}"
        )
    return uri


def get_mongo_client(uri: str) -> tuple[MongoClient, str]:
    db_name = database_name_from_uri(uri)
    client = MongoClient(uri, serverSelectionTimeoutMS=10_000)
    client.admin.command("ping")
    return client, db_name


def mongo_prepare_db(scenario_dir: Path, cfg: ScenarioConfig) -> None:
    uri = mongo_uri_for_scenario(scenario_dir, cfg)
    client, db_name = get_mongo_client(uri)
    try:
        db = client[db_name]
        db["transactions"].delete_many({})
        db["users"].delete_many({})
        raw = USERS_SEED_FILE.read_text(encoding="utf-8")
        users_docs: list[Any] = json_util.loads(raw)
        if not isinstance(users_docs, list):
            raise ValueError("users-collection.json deve ser um array JSON")
        if users_docs:
            db["users"].insert_many(users_docs)
    finally:
        client.close()


def prometheus_data_volume_name(scenario_folder_name: str, execution_index: int) -> str:
    """
    Nome estável do volume Docker para TSDB do Prometheus nesta pasta/rodada.

    Ex.: pasta ``scenario-2``, rodada 1 → ``scenario_2_1_execution``; rodada 2 → ``scenario_2_2_execution``.
    Pastas que não casam ``scenario-<número>`` usam o nome da pasta normalizado.
    """
    m = re.fullmatch(r"scenario-(\d+)", scenario_folder_name, flags=re.IGNORECASE)
    if m:
        key = m.group(1)
    else:
        key = re.sub(r"[^\w]+", "_", scenario_folder_name).strip("_").lower() or "unknown"
    return f"scenario_{key}_{int(execution_index)}_execution"


def scenario_compose_persists_prometheus_tsdb(scenario_dir: Path) -> bool:
    """True se ``docker-compose.yml`` do cenário monta o volume TSDB externo do Prometheus."""
    p = scenario_dir / "docker-compose.yml"
    if not p.is_file():
        return False
    try:
        return "prometheus_tsdb:/prometheus" in p.read_text(encoding="utf-8")
    except OSError:
        return False


def ensure_docker_volume_exists(name: str) -> None:
    """Garante volume nomeado (``docker volume create`` se ainda não existir)."""
    r = subprocess.run(
        ["docker", "volume", "inspect", name],
        capture_output=True,
        text=True,
        timeout=30,
    )
    if r.returncode == 0:
        print(f"Volume Prometheus (existente): {name}")
        return
    subprocess.run(["docker", "volume", "create", name], check=True, timeout=60)
    print(f"Volume Prometheus criado: {name}")


def compose_base_cmd() -> list[str]:
    if shutil.which("docker"):
        r = subprocess.run(
            ["docker", "compose", "version"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        if r.returncode == 0:
            return ["docker", "compose"]
    exe = shutil.which("docker-compose")
    if exe:
        return [exe]
    raise RuntimeError("docker compose ou docker-compose não encontrado no PATH")


def _compose_subprocess_env(extra: dict[str, str] | None) -> dict[str, str]:
    env = os.environ.copy()
    if extra:
        env.update(extra)
    return env


def docker_compose_up(scenario_dir: Path, *, compose_env: dict[str, str] | None = None) -> None:
    cmd = [*compose_base_cmd(), "up", "-d", "--build"]
    print(f"Subindo stack: {' '.join(cmd)} (cwd={scenario_dir})")
    subprocess.run(cmd, cwd=scenario_dir, check=True, env=_compose_subprocess_env(compose_env))


def rm_containers(scenario_dir: Path, *, compose_env: dict[str, str] | None = None) -> None:
    """Para e remove containers, rede default do projeto e volumes não externos do compose.

    Volumes Prometheus marcados como ``external: true`` (TSDB por execução) permanecem no host.
    """
    cmd = [*compose_base_cmd(), "down", "--remove-orphans", "-v"]
    print(f"Derrubando stack: {' '.join(cmd)} (cwd={scenario_dir})")
    try:
        subprocess.run(cmd, cwd=scenario_dir, check=True, env=_compose_subprocess_env(compose_env))
    except subprocess.CalledProcessError as e:
        print(f"Falha ao derrubar stack: {e}", file=sys.stderr)


def compose_network_name(scenario_dir: Path) -> str:
    project = os.environ.get("COMPOSE_PROJECT_NAME") or scenario_dir.name
    return f"{project}_monitoring"


def tcp_port_open(host: str, port: int, timeout_s: float = 1.0) -> bool:
    try:
        with socket.create_connection((host, port), timeout=timeout_s):
            return True
    except OSError:
        return False


def wait_for_ports(
    deadline: float,
    checks: tuple[tuple[str, int], ...] | None = None,
) -> None:
    port_checks = checks if checks is not None else DEFAULT_PORT_CHECKS
    while time.time() < deadline:
        ok = all(tcp_port_open("127.0.0.1", port) for _, port in port_checks)
        if ok:
            return
        time.sleep(PORT_RETRY_INTERVAL_S)
    missing = [name for name, port in port_checks if not tcp_port_open("127.0.0.1", port)]
    raise TimeoutError(f"Portas não responderam a tempo: {missing}")


def wait_for_app_health(deadline: float, health_url: str) -> None:
    while time.time() < deadline:
        try:
            r = requests.get(health_url, timeout=10, verify=os.environ.get("HEALTH_TLS_VERIFY", "true").lower() != "false")
            if r.status_code == 200:
                body = r.json()
                if body.get("status") == "UP":
                    return
        except requests.RequestException:
            pass
        time.sleep(PORT_RETRY_INTERVAL_S)
    raise TimeoutError(f"Health check não passou: {health_url}")


def run_k6(scenario_dir: Path, cfg: ScenarioConfig) -> tuple[int, int, int]:
    """Executa k6 com script em tests/ e saídas em scenario_dir. Retorna (start, end, exit).

    BASE_URL vem de config.applicationUri. Sem docker compose, o container k6 usa rede
    padrão do Docker (acesso à API pública). Com compose, usa a rede do cenário.
    """
    tr = TESTS_ROOT.resolve()
    script_name = Path(cfg.testFile).name
    script_path = tr / script_name
    if not script_path.is_file():
        raise FileNotFoundError(f"Script k6 não encontrado: {script_path}")

    sdir = scenario_dir.resolve()
    for fname in ("k6-result.json", "k6-summary.json"):
        p = sdir / fname
        if p.exists():
            p.unlink()

    cmd: list[str] = ["docker", "run", "--rm", "-i"]
    if cfg.runDockerCompose and not cfg.externalApplication:
        net = compose_network_name(scenario_dir)
        cmd.extend(["--network", net])
        print(f"Executando k6 (rede {net}, BASE_URL={cfg.applicationUri})...")
    else:
        print(f"Executando k6 (rede default do Docker, BASE_URL={cfg.applicationUri})...")

    cmd.extend(
        [
            "-v",
            f"{tr}:/scripts:ro",
            "-v",
            f"{sdir}:/out",
            "-e",
            f"BASE_URL={cfg.applicationUri}",
            "grafana/k6",
            "run",
            "--summary-export",
            "/out/k6-summary.json",
            "--out",
            "json=/out/k6-result.json",
            f"/scripts/{script_name}",
        ]
    )

    start_time = int(time.time())
    proc = subprocess.run(cmd, check=False)
    end_time = int(time.time())
    code = int(proc.returncode or 0)
    if code != 0:
        print(
            f"Aviso: k6 terminou com código {code} (ex.: thresholds cruzados). "
            "Seguindo com capturas, Prometheus e planilha; próximos cenários não são cancelados.",
            file=sys.stderr,
        )
    return start_time, end_time, code


def parse_k6_results(k6_json: Path) -> dict[str, Any]:
    metrics: dict[str, Any] = {"rps_avg": 0, "latency_p95": 0, "latency_p99": 0, "errors": 0}
    if not k6_json.is_file():
        return metrics
    with open(k6_json, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            data = json.loads(line)
            if data.get("type") != "Point":
                continue
            metric = data.get("metric")
            if metric == "http_reqs":
                metrics["rps_avg"] += 1
            if metric == "http_req_duration":
                value = data["data"]["value"]
                metrics.setdefault("latencies", []).append(value)
    if "latencies" in metrics and metrics["latencies"]:
        lat = sorted(metrics["latencies"])
        i95 = min(int(len(lat) * 0.95), len(lat) - 1)
        i99 = min(int(len(lat) * 0.99), len(lat) - 1)
        metrics["latency_p95"] = lat[i95]
        metrics["latency_p99"] = lat[i99]
    return metrics


def _flatten_summary_checks(root: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    stack: list[dict[str, Any]] = [root]
    while stack:
        g = stack.pop()
        for c in g.get("checks") or []:
            if isinstance(c, dict):
                out.append(c)
        for sub in g.get("groups") or []:
            if isinstance(sub, dict):
                stack.append(sub)
    return out


def parse_k6_summary(summary_path: Path) -> dict[str, Any]:
    """Lê k6-summary.json (--summary-export) para percentis e checks."""
    if not summary_path.is_file():
        return {}
    try:
        with open(summary_path, encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}
    metrics = data.get("metrics") or {}
    dur = metrics.get("http_req_duration") or {}
    vals = dur.get("values") or {}
    p95 = vals.get("p(95)")
    p99 = vals.get("p(99)")
    failed = metrics.get("http_req_failed") or {}
    fr_vals = failed.get("values") or {}
    fail_rate = fr_vals.get("rate")
    root = data.get("root_group") or {}
    checks = _flatten_summary_checks(root)
    return {
        "p95_ms": p95,
        "p99_ms": p99,
        "http_fail_rate": fail_rate,
        "checks": checks,
    }


def compute_sla_flags(summary: dict[str, Any], infra_data: dict[str, Any]) -> dict[str, bool]:
    """Indicadores da pesquisa: latência, falhas, checks de negócio, CPU e memória (Prometheus)."""
    p95 = summary.get("p95_ms")
    p99 = summary.get("p99_ms")
    fr = summary.get("http_fail_rate")
    k6_p95_ok = p95 is not None and float(p95) < 1000
    k6_p99_ok = p99 is not None and float(p99) < 2000
    k6_fail_ok = fr is not None and float(fr) < 0.05

    checks = summary.get("checks") or []
    biz_ok = bool(checks) and all(int(c.get("fails") or 0) == 0 for c in checks)

    cpu_max = infra_data.get("cpu_max")
    cpu_ok = cpu_max is not None and float(cpu_max) < 80.0

    mem_ratio_max = infra_data.get("memory_pressure_max")
    mem_ok = mem_ratio_max is not None and float(mem_ratio_max) <= 0.70 + 1e-9

    return {
        "SLA K6 P95 < 1s": k6_p95_ok,
        "SLA K6 P99 < 2s": k6_p99_ok,
        "SLA K6 taxa falha < 5%": k6_fail_ok,
        "SLA checks negócio (todos passando)": biz_ok,
        "SLA Prometheus CPU < 80%": cpu_ok,
        "SLA Prometheus memória < 70% limite": mem_ok,
    }


def query_prometheus(
    query: str,
    start: int,
    end: int,
    step: str = "5s",
    *,
    base_url: str | None = None,
) -> tuple[Any, Any]:
    base = (base_url or PROMETHEUS_URL).rstrip("/")
    url = f"{base}/api/v1/query_range"
    verify = os.environ.get("PROMETHEUS_TLS_VERIFY", "true").lower() != "false"
    response = requests.get(
        url,
        params={"query": query, "start": start, "end": end, "step": step},
        timeout=60,
        verify=verify,
    )
    response.raise_for_status()
    data = response.json()
    if data.get("status") != "success":
        return None, None
    results = data.get("data", {}).get("result") or []
    values: list[float] = []
    for result in results:
        for v in result.get("values") or []:
            try:
                values.append(float(v[1]))
            except (TypeError, ValueError, IndexError):
                pass
    if not values:
        return None, None
    return min(values), max(values)


def collect_metrics(
    start: int,
    end: int,
    *,
    prometheus_base_url: str | None = None,
) -> dict[str, Any]:
    queries = {
        "cpu": 'rate(container_cpu_usage_seconds_total{name="transaction-api"}[1m]) * 100',
        "memory": 'container_memory_usage_bytes{name="transaction-api"} / 1024 / 1024 / 1024',
        "memory_pressure": 'container_memory_usage_bytes{name="transaction-api"} / container_spec_memory_limit_bytes{name="transaction-api"}',
        "latency_p95": "histogram_quantile(0.95, sum(rate(http_server_requests_seconds_bucket[1m])) by (le)) * 1000",
        "rps": "sum(rate(http_server_requests_seconds_count[1m]))",
        "threads_busy": "tomcat_threads_busy_threads",
        "threads_max": "tomcat_threads_config_max_threads",
        "jvm_threads": "jvm_threads_live_threads",
    }
    metrics: dict[str, Any] = {}
    for name, query in queries.items():
        try:
            min_val, max_val = query_prometheus(
                query, start, end, base_url=prometheus_base_url
            )
        except (requests.RequestException, KeyError, ValueError) as e:
            print(f"Aviso: métrica Prometheus '{name}' indisponível: {e}", file=sys.stderr)
            min_val, max_val = None, None
        metrics[f"{name}_min"] = min_val
        metrics[f"{name}_max"] = max_val
    return metrics


def _iter_dashboard_panels(dashboard: dict[str, Any]) -> list[tuple[int, str]]:
    """Painéis com id (ignora só o tipo row, percorre painéis aninhados)."""

    out: list[tuple[int, str]] = []

    def walk(panels: list[Any] | None) -> None:
        for panel in panels or []:
            if panel.get("type") == "row":
                walk(panel.get("panels"))
                continue
            pid = panel.get("id")
            if pid is None:
                continue
            title = str(panel.get("title") or f"panel-{pid}")
            out.append((int(pid), title))

    walk(dashboard.get("panels"))
    out.sort(key=lambda t: t[0])
    return out


def _safe_filename_fragment(s: str, max_len: int = 60) -> str:
    s = re.sub(r"[^\w\s\-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_]+", "-", s.strip()).strip("-")
    return (s or "panel")[:max_len]


def _grafana_dashboard_slug(uid: str, base_url: str, user: str, password: str) -> str:
    url = f"{base_url}/api/dashboards/uid/{uid}"
    r = requests.get(url, auth=(user, password), timeout=30)
    r.raise_for_status()
    meta = r.json().get("meta") or {}
    return str(meta.get("slug") or uid)


def capture_grafana_dashboard_panels(
    start_s: int,
    end_s: int,
    scenario_dir: Path,
    dashboard_json_path: Path,
    *,
    skip: bool = False,
) -> Path | None:
    """
    Após o k6, abre cada painel do dashboard no modo /d-solo/ (Grafana) e salva PNG.

    Não depende do plugin Grafana Image Renderer: usa Playwright (Chromium).
    Credenciais: GRAFANA_USER / GRAFANA_PASSWORD (padrão admin/admin).
    """
    if skip:
        return None
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print(
            "Playwright não instalado; pulando capturas Grafana. "
            "Instale: pip install playwright && playwright install chromium",
            file=sys.stderr,
        )
        return None

    if not dashboard_json_path.is_file():
        print(f"Dashboard JSON não encontrado: {dashboard_json_path}", file=sys.stderr)
        return None

    with open(dashboard_json_path, encoding="utf-8") as f:
        dashboard = json.load(f)
    uid = str(dashboard.get("uid") or "")
    if not uid:
        print("Dashboard JSON sem campo uid.", file=sys.stderr)
        return None

    panels = _iter_dashboard_panels(dashboard)
    if not panels:
        print("Nenhum painel encontrado no JSON do dashboard.", file=sys.stderr)
        return None

    user = os.environ.get("GRAFANA_USER", "admin")
    password = os.environ.get("GRAFANA_PASSWORD", "admin")
    try:
        slug = _grafana_dashboard_slug(uid, GRAFANA_URL, user, password)
    except requests.RequestException as e:
        print(
            f"Não foi possível obter slug do dashboard via API Grafana ({e}); "
            "verifique GRAFANA_URL, usuário e senha.",
            file=sys.stderr,
        )
        return None

    pad_ms = 60_000
    from_ms = max(0, start_s * 1000 - pad_ms)
    to_ms = end_s * 1000 + pad_ms

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out_dir = scenario_dir / "grafana-captures" / stamp
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Capturando {len(panels)} painel(is) Grafana em {out_dir} ...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-dev-shm-usage", "--no-sandbox"],
        )
        try:
            context = browser.new_context(
                viewport={"width": 1600, "height": 1000},
                ignore_https_errors=True,
            )
            page = context.new_page()
            page.set_default_timeout(90_000)

            login_url = f"{GRAFANA_URL}/login"
            page.goto(login_url, wait_until="domcontentloaded")
            page.locator('input[name="user"]').fill(user)
            page.locator('input[name="password"]').fill(password)
            page.locator('button[type="submit"]').click()
            page.wait_for_load_state("networkidle")

            for panel_id, title in panels:
                fname = f"panel-{panel_id:02d}-{_safe_filename_fragment(title)}.png"
                dest = out_dir / fname
                solo = (
                    f"{GRAFANA_URL}/d-solo/{uid}/{slug}"
                    f"?orgId=1&from={from_ms}&to={to_ms}&panelId={panel_id}"
                    f"&theme=light&kiosk"
                )
                try:
                    page.goto(solo, wait_until="networkidle")
                    page.wait_for_timeout(2500)
                    page.screenshot(path=str(dest), full_page=True)
                    print(f"  OK {fname}")
                except Exception as ex:
                    print(f"  Falha painel {panel_id} ({title}): {ex}", file=sys.stderr)
        finally:
            browser.close()

    return out_dir


def export_to_excel(
    k6_data: dict[str, Any],
    infra_data: dict[str, Any],
    cfg: ScenarioConfig,
    scenario_folder_name: str,
    k6_exit_code: int = 0,
    sla_flags: dict[str, bool] | None = None,
    *,
    test_start_epoch: int | None = None,
    test_end_epoch: int | None = None,
) -> None:
    db_label = "interno" if cfg.databaseLocation == "internal" else "externo (cloud)"
    sla_flags = sla_flags or {}
    formatted_data: dict[str, Any] = {
        "Pasta do cenário": scenario_folder_name,
        "Cenario": cfg.scenario,
        "Nome do Teste": cfg.testFile,
        "applicationUri": cfg.applicationUri,
        "externalApplication": cfg.externalApplication,
        "prometheusUri": cfg.prometheusUri or "",
        "Descrição": cfg.description,
        "CPU Alocada": cfg.cpus,
        "Memory Alocada": cfg.memory,
        "Banco de Dados": db_label,
        "Data e hora de início": (
            format_datetime_local_from_epoch(test_start_epoch)
            if test_start_epoch is not None
            else ""
        ),
        "Data e hora de fim": (
            format_datetime_local_from_epoch(test_end_epoch)
            if test_end_epoch is not None
            else ""
        ),
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Código saída k6": k6_exit_code,
        "RPS Médio": k6_data.get("rps_avg"),
        "Latência P95 (ms)": k6_data.get("latency_p95"),
        "Latência P99 (ms)": k6_data.get("latency_p99"),
        "Erros": k6_data.get("errors"),
        "CPU Min (%)": infra_data.get("cpu_min"),
        "CPU Max (%)": infra_data.get("cpu_max"),
        "Memória Min (GB)": infra_data.get("memory_min"),
        "Memória Max (GB)": infra_data.get("memory_max"),
        "Pressão memória max (0-1)": infra_data.get("memory_pressure_max"),
        "Latência Infra Min (ms)": infra_data.get("latency_p95_min"),
        "Latência Infra Max (ms)": infra_data.get("latency_p95_max"),
        "RPS Min": infra_data.get("rps_min"),
        "RPS Max": infra_data.get("rps_max"),
        "Threads Busy Min": infra_data.get("threads_busy_min"),
        "Threads Busy Max": infra_data.get("threads_busy_max"),
        "Threads Max Config": infra_data.get("threads_max_max"),
        "JVM Threads Min": infra_data.get("jvm_threads_min"),
        "JVM Threads Max": infra_data.get("jvm_threads_max"),
    }
    for k, v in sla_flags.items():
        formatted_data[k] = v
    df = pd.DataFrame([formatted_data])
    if OUTPUT_FILE.exists():
        existing = pd.read_excel(OUTPUT_FILE)
        df = pd.concat([existing, df], ignore_index=True)
    df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.column).value
            if cell.value is None or not header:
                continue
            if "(%)" in str(header):
                if isinstance(cell.value, (int, float)) and cell.value > 1:
                    cell.value = cell.value / 100
                cell.number_format = "0.00%"
            elif "(ms)" in str(header):
                cell.number_format = "0"
            elif "(GB)" in str(header):
                cell.number_format = "0.00"
    wb.save(OUTPUT_FILE)
    print(f"Planilha atualizada: {OUTPUT_FILE}")


def prompt_before_docker_teardown(scenario_name: str, *, pause: bool) -> None:
    """
    Espera Enter no terminal antes de docker compose down, para o operador
    analisar Grafana/Prometheus com os containers ainda no ar.
    """
    if not pause:
        return
    if not sys.stdin.isatty():
        print(
            "Aviso: --pause-before-docker-down ignorado (stdin não é um terminal interativo).",
            file=sys.stderr,
        )
        return
    print(
        f"\n{'=' * 60}\n"
        f"Pausa após o cenário «{scenario_name}»\n"
        f"Grafana: {GRAFANA_URL}  |  Prometheus local (se aplicável): {PROMETHEUS_URL}\n"
        f"Analise as métricas e pressione ENTER para executar docker compose down e continuar\n"
        f"(próximo cenário ou fim do fluxo).\n"
        f"{'=' * 60}\n",
        flush=True,
    )
    try:
        input()
    except EOFError:
        print("(EOF recebido — seguindo com encerramento dos containers.)", file=sys.stderr)


def run_scenario(
    scenario_dir: Path,
    *,
    execution_index: int = 1,
    skip_grafana_captures: bool = False,
    pause_before_docker_down: bool = False,
) -> None:
    name = scenario_dir.name
    print(f"\n======== Cenário: {name} (rodada {execution_index}) ========")
    cfg: ScenarioConfig | None = None
    deadline = time.time() + COMPOSE_SERVICE_WAIT_S
    compose_env: dict[str, str] | None = None
    try:
        cfg = load_scenario_config(scenario_dir)
        health_url = resolve_health_check_url(cfg)
        if cfg.runDockerCompose:
            if cfg.externalApplication:
                if not cfg.prometheusUri:
                    print(
                        "Aviso: externalApplication=true sem prometheusUri; "
                        "métricas na planilha podem falhar e o datasource Grafana não será gerado.",
                        file=sys.stderr,
                    )
                else:
                    print('')
                    # write_grafana_prometheus_datasource(scenario_dir, cfg.prometheusUri)
            if scenario_compose_persists_prometheus_tsdb(scenario_dir):
                vol = prometheus_data_volume_name(name, execution_index)
                ensure_docker_volume_exists(vol)
                compose_env = {"PROMETHEUS_DATA_VOLUME_NAME": vol}
                print(
                    f"TSDB Prometheus será persistido no volume Docker «{vol}» "
                    f"(ver tests/docker-compose.metrics-view.yml para reabrir com Grafana)."
                )
            docker_compose_up(scenario_dir, compose_env=compose_env)
            wait_for_ports(deadline, port_checks_for_config(cfg))
        else:
            print("runDockerCompose=false: pulando docker compose e verificação de portas locais.")
        wait_for_app_health(deadline, health_url)
        mongo_prepare_db(scenario_dir, cfg)
        start, end, k6_exit = run_k6(scenario_dir, cfg)
        skip_gf = skip_grafana_captures or (not cfg.runDockerCompose)
        dashboard_json = grafana_dashboard_json_for_scenario(scenario_dir)
        if not skip_gf:
            print(f"Capturas Grafana: dashboard JSON = {dashboard_json}")
        capture_grafana_dashboard_panels(
            start,
            end,
            scenario_dir,
            dashboard_json,
            skip=skip_gf,
        )
        k6_path = scenario_dir / "k6-result.json"
        summary_path = scenario_dir / "k6-summary.json"
        summary = parse_k6_summary(summary_path)
        k6_data = parse_k6_results(k6_path)
        if summary.get("p95_ms") is not None:
            k6_data["latency_p95"] = float(summary["p95_ms"])
        if summary.get("p99_ms") is not None:
            k6_data["latency_p99"] = float(summary["p99_ms"])
        print("Coletando métricas do Prometheus...")
        prom_base = cfg.prometheusUri if cfg.externalApplication else None
        infra_data = collect_metrics(start, end, prometheus_base_url=prom_base)
        sla_flags = compute_sla_flags(summary, infra_data)
        export_to_excel(
            k6_data,
            infra_data,
            cfg,
            name,
            k6_exit_code=k6_exit,
            sla_flags=sla_flags,
            test_start_epoch=start,
            test_end_epoch=end,
        )
        if k6_exit != 0:
            print(f"Cenário {name} concluído (k6 exit {k6_exit}; fluxo seguiu até o fim).")
        else:
            print(f"Cenário {name} concluído com sucesso.")
    finally:
        if cfg is not None and cfg.runDockerCompose:
            prompt_before_docker_teardown(name, pause=pause_before_docker_down)
            rm_containers(scenario_dir, compose_env=compose_env)


def main() -> int:
    parser = argparse.ArgumentParser(description="Roda cenários de carga e gera result.xlsx")
    parser.add_argument(
        "--tests-root",
        type=Path,
        default=TESTS_ROOT,
        help="Diretório tests (contém scenario-*, users-collection.json)",
    )
    parser.add_argument(
        "--scenario",
        type=str,
        default="",
        metavar="PASTA",
        help="Nome exato da pasta do cenário (ex.: scenario-2). Sem isso, roda todos.",
    )
    parser.add_argument(
        "--only",
        type=str,
        default="",
        help="Regex no nome da pasta (ex.: ^scenario-1$). Ignorado se --scenario for usado.",
    )
    parser.add_argument(
        "--skip-scenario",
        action="append",
        default=None,
        metavar="PASTA",
        help="Nome exato da pasta a não executar (oposto de --scenario). Pode repetir para pular vários.",
    )
    parser.add_argument(
        "--repeat",
        type=int,
        default=1,
        metavar="N",
        help="Quantas vezes repetir o fluxo completo (todos os cenários da lista). Padrão: 1.",
    )
    parser.add_argument(
        "--skip-grafana-captures",
        action="store_true",
        help="Não gera PNGs dos painéis do Grafana após o k6",
    )
    parser.add_argument(
        "--pause-before-docker-down",
        action="store_true",
        help=(
            "Após cada cenário (k6, capturas Grafana, planilha), espera ENTER antes de "
            "docker compose down, para analisar o Grafana com os containers ainda ativos."
        ),
    )
    args = parser.parse_args()
    tests_root = args.tests_root.resolve()
    if args.repeat < 1:
        print("--repeat deve ser >= 1.", file=sys.stderr)
        return 1

    if not USERS_SEED_FILE.is_file():
        print(f"Arquivo de seed não encontrado: {USERS_SEED_FILE}", file=sys.stderr)
        return 1

    all_scenarios = discover_scenarios(tests_root)
    scenarios = all_scenarios
    if args.scenario.strip():
        want = args.scenario.strip()
        scenarios = [s for s in all_scenarios if s.name == want]
        if not scenarios:
            print(
                f"Cenário '{want}' não encontrado em {tests_root}. "
                f"Pastas válidas: {[p.name for p in all_scenarios]}",
                file=sys.stderr,
            )
            return 1
    elif args.only:
        pattern = re.compile(args.only)
        scenarios = [s for s in all_scenarios if pattern.search(s.name)]

    skip_names = {s.strip() for s in (args.skip_scenario or []) if s and str(s).strip()}
    if skip_names:
        in_queue = {s.name for s in scenarios}
        skipped = sorted(skip_names & in_queue)
        unknown_skip = sorted(skip_names - in_queue)
        scenarios = [s for s in scenarios if s.name not in skip_names]
        if skipped:
            print(f"Cenários excluídos (--skip-scenario): {skipped}")
        if unknown_skip:
            print(
                f"Aviso: --skip-scenario ignorado (não estavam na fila): {unknown_skip}",
                file=sys.stderr,
            )

    if not scenarios:
        print(
            f"Nenhum cenário a executar em {tests_root} (filtros resultaram em lista vazia).",
            file=sys.stderr,
        )
        return 1

    print(f"Cenários a executar: {[s.name for s in scenarios]}")
    if args.repeat > 1:
        print(f"Repetições do fluxo completo: {args.repeat}")

    flow_start = time.perf_counter()
    for run_idx in range(1, args.repeat + 1):
        if args.repeat > 1:
            print(f"\n{'='*60}\nRodada {run_idx} de {args.repeat}\n{'='*60}")

        for si, scenario_dir in enumerate(scenarios):
            cleanup_env: dict[str, str] | None = None
            if scenario_compose_persists_prometheus_tsdb(scenario_dir):
                cleanup_env = {
                    "PROMETHEUS_DATA_VOLUME_NAME": prometheus_data_volume_name(
                        scenario_dir.name, run_idx
                    )
                }
            try:
                run_scenario(
                    scenario_dir,
                    execution_index=run_idx,
                    skip_grafana_captures=args.skip_grafana_captures,
                    pause_before_docker_down=args.pause_before_docker_down,
                )
            except subprocess.CalledProcessError as e:
                print(f"Erro de subprocesso no cenário {scenario_dir.name}: {e}", file=sys.stderr)
                try:
                    c = load_scenario_config(scenario_dir)
                    if c.runDockerCompose:
                        rm_containers(scenario_dir, compose_env=cleanup_env)
                except Exception:
                    if (scenario_dir / "docker-compose.yml").is_file():
                        rm_containers(scenario_dir, compose_env=cleanup_env)
                return 1
            except Exception as e:
                print(f"Falha no cenário {scenario_dir.name}: {e}", file=sys.stderr)
                try:
                    c = load_scenario_config(scenario_dir)
                    if c.runDockerCompose:
                        rm_containers(scenario_dir, compose_env=cleanup_env)
                except Exception:
                    if (scenario_dir / "docker-compose.yml").is_file():
                        rm_containers(scenario_dir, compose_env=cleanup_env)
                return 1

            last_scenario = si == len(scenarios) - 1
            last_round = run_idx == args.repeat
            if not (last_scenario and last_round):
                print("\nAguardando para iniciar próximo cenário...\n")
                time.sleep(60)

    elapsed = time.perf_counter() - flow_start
    print(f"\nTempo total do fluxo (início → fim): {format_duration_hms(elapsed)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
